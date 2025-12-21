"""Excel report generation and PORT export."""

import pandas as pd
from pathlib import Path
from typing import Optional
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows

from .database import get_latest_positions
from .analytics import (
    get_consolidated_positions,
    get_portfolio_summary,
    get_segmentation_by_security_type,
    get_segmentation_by_payment_rank,
    get_concentration_by_issuer,
    get_concentration_by_country,
    get_concentration_by_rating,
    get_concentration_by_currency,
    get_concentration_by_maturity,
    check_concentration_alerts,
    get_duplicate_holdings,
)


# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
MONEY_FORMAT = '#,##0.00'
PCT_FORMAT = '0.00%'


def _style_header_row(ws, row_num: int, col_count: int):
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _add_dataframe_to_sheet(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    """Add a DataFrame to a worksheet with formatting."""
    if df.empty:
        ws.cell(row=start_row, column=start_col, value="No data available")
        return start_row + 1

    # Headers
    for col_idx, col_name in enumerate(df.columns, start_col):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, start_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER

            # Format numbers
            if isinstance(value, float):
                if "pct" in df.columns[col_idx - start_col].lower():
                    cell.number_format = '0.00'
                else:
                    cell.number_format = MONEY_FORMAT

    return start_row + len(df) + 1


def generate_consolidation_report(
    output_path: Path,
    db_path: Optional[Path] = None
) -> int:
    """
    Generate consolidation report Excel file.

    Returns the number of positions included.
    """
    positions = get_consolidated_positions(db_path)

    if positions.empty:
        print("No positions to report.")
        return 0

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary = get_portfolio_summary(db_path)
    ws_summary.cell(row=1, column=1, value="Portfolio Consolidation Report").font = Font(bold=True, size=14)
    ws_summary.cell(row=2, column=1, value=f"Generated: {date.today().isoformat()}")

    ws_summary.cell(row=4, column=1, value="Portfolio Summary").font = Font(bold=True, size=12)
    ws_summary.cell(row=5, column=1, value="Total Positions:")
    ws_summary.cell(row=5, column=2, value=summary["total_positions"])
    ws_summary.cell(row=6, column=1, value="Unique Securities:")
    ws_summary.cell(row=6, column=2, value=summary["unique_securities"])
    ws_summary.cell(row=7, column=1, value="Banks:")
    ws_summary.cell(row=7, column=2, value=summary["bank_count"])
    ws_summary.cell(row=8, column=1, value="Total Value (USD):")
    ws_summary.cell(row=8, column=2, value=summary["total_value_usd"]).number_format = MONEY_FORMAT

    # Positions sheet
    ws_pos = wb.create_sheet("Positions")

    # Select columns for report
    report_cols = [
        "bank_code", "bank_name", "isin", "security_name", "issuer",
        "security_type", "payment_rank", "currency", "face_value",
        "market_value", "market_value_usd", "exposure_pct",
        "coupon_rate", "security_maturity", "rating_sp", "rating_moody",
        "modified_duration", "yield_to_maturity", "issuer_country"
    ]

    # Filter to available columns
    available_cols = [c for c in report_cols if c in positions.columns]
    report_df = positions[available_cols].copy()

    # Rename columns for readability
    col_names = {
        "bank_code": "Bank",
        "bank_name": "Bank Name",
        "isin": "ISIN",
        "security_name": "Security",
        "issuer": "Issuer",
        "security_type": "Type",
        "payment_rank": "Seniority",
        "currency": "CCY",
        "face_value": "Face Value",
        "market_value": "Market Value",
        "market_value_usd": "Value (USD)",
        "exposure_pct": "% of Portfolio",
        "coupon_rate": "Coupon",
        "security_maturity": "Maturity",
        "rating_sp": "S&P",
        "rating_moody": "Moody's",
        "modified_duration": "Duration",
        "yield_to_maturity": "YTM",
        "issuer_country": "Country"
    }
    report_df = report_df.rename(columns=col_names)

    _add_dataframe_to_sheet(ws_pos, report_df)

    # Auto-adjust column widths
    for col_idx, col in enumerate(report_df.columns, 1):
        max_width = max(len(str(col)), report_df[col].astype(str).str.len().max())
        ws_pos.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"[0:2]].width = min(max_width + 2, 30)

    # By Bank sheet
    ws_bank = wb.create_sheet("By Bank")
    by_bank = positions.groupby(["bank_code", "bank_name"]).agg({
        "isin": "count",
        "market_value_usd": "sum"
    }).reset_index()
    by_bank.columns = ["Bank Code", "Bank Name", "Positions", "Value (USD)"]
    _add_dataframe_to_sheet(ws_bank, by_bank)

    # By Security Type sheet
    ws_type = wb.create_sheet("By Type")
    seg_type = get_segmentation_by_security_type(db_path)
    if not seg_type.empty:
        seg_type.columns = ["Security Type", "Positions", "Value (USD)", "Avg Duration", "Avg YTM", "% of Portfolio"]
        _add_dataframe_to_sheet(ws_type, seg_type)

    # Duplicates sheet (same security at multiple banks)
    ws_dup = wb.create_sheet("Duplicates")
    duplicates = get_duplicate_holdings(db_path)
    if not duplicates.empty:
        dup_report = duplicates[["isin", "security_name", "issuer", "bank_code", "market_value_usd"]].copy()
        dup_report["bank_code"] = dup_report["bank_code"].apply(lambda x: ", ".join(x))
        dup_report.columns = ["ISIN", "Security", "Issuer", "Banks", "Total Value (USD)"]
        _add_dataframe_to_sheet(ws_dup, dup_report)
    else:
        ws_dup.cell(row=1, column=1, value="No duplicate holdings found")

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return len(positions)


def generate_concentration_report(
    output_path: Path,
    db_path: Optional[Path] = None
) -> int:
    """
    Generate concentration risk report Excel file.

    Returns the number of alerts found.
    """
    wb = Workbook()

    # Alerts sheet
    ws_alerts = wb.active
    ws_alerts.title = "Alerts"

    ws_alerts.cell(row=1, column=1, value="Concentration Risk Report").font = Font(bold=True, size=14)
    ws_alerts.cell(row=2, column=1, value=f"Generated: {date.today().isoformat()}")

    alerts = check_concentration_alerts(db_path)

    if alerts:
        ws_alerts.cell(row=4, column=1, value=f"Found {len(alerts)} concentration alerts").font = Font(bold=True)

        alert_data = []
        for alert in alerts:
            alert_data.append({
                "Category": alert.category,
                "Name": alert.name,
                "Exposure (USD)": alert.exposure_usd,
                "Exposure %": alert.exposure_pct,
                "Alert Type": alert.threshold_type.upper(),
                "Threshold %": alert.threshold_value
            })

        alert_df = pd.DataFrame(alert_data)
        last_row = _add_dataframe_to_sheet(ws_alerts, alert_df, start_row=6)

        # Color code by alert type
        for row_idx in range(7, last_row):
            alert_type_cell = ws_alerts.cell(row=row_idx, column=5)
            if alert_type_cell.value == "CRITICAL":
                for col in range(1, 7):
                    ws_alerts.cell(row=row_idx, column=col).fill = CRITICAL_FILL
            elif alert_type_cell.value == "WARNING":
                for col in range(1, 7):
                    ws_alerts.cell(row=row_idx, column=col).fill = WARNING_FILL
    else:
        ws_alerts.cell(row=4, column=1, value="No concentration alerts - portfolio is well diversified").font = Font(color="006400")

    # By Issuer sheet
    ws_issuer = wb.create_sheet("By Issuer")
    issuer_conc = get_concentration_by_issuer(db_path)
    if not issuer_conc.empty:
        issuer_conc.columns = ["Issuer", "Country", "Securities", "Value (USD)", "% of Portfolio"]
        _add_dataframe_to_sheet(ws_issuer, issuer_conc.head(50))  # Top 50

    # By Country sheet
    ws_country = wb.create_sheet("By Country")
    country_conc = get_concentration_by_country(db_path)
    if not country_conc.empty:
        country_conc.columns = ["Country", "Securities", "Issuers", "Value (USD)", "% of Portfolio"]
        _add_dataframe_to_sheet(ws_country, country_conc)

    # By Rating sheet
    ws_rating = wb.create_sheet("By Rating")
    rating_conc = get_concentration_by_rating(db_path)
    if not rating_conc.empty:
        rating_conc.columns = ["Rating Bucket", "Securities", "Value (USD)", "% of Portfolio"]
        _add_dataframe_to_sheet(ws_rating, rating_conc)

    # By Currency sheet
    ws_ccy = wb.create_sheet("By Currency")
    ccy_conc = get_concentration_by_currency(db_path)
    if not ccy_conc.empty:
        ccy_conc.columns = ["Currency", "Securities", "Value (USD)", "% of Portfolio"]
        _add_dataframe_to_sheet(ws_ccy, ccy_conc)

    # By Maturity sheet
    ws_mat = wb.create_sheet("By Maturity")
    mat_conc = get_concentration_by_maturity(db_path)
    if not mat_conc.empty:
        mat_conc.columns = ["Maturity Bucket", "Securities", "Value (USD)", "Avg Duration", "% of Portfolio"]
        _add_dataframe_to_sheet(ws_mat, mat_conc)

    # By Security Type sheet
    ws_type = wb.create_sheet("By Security Type")
    type_conc = get_segmentation_by_security_type(db_path)
    if not type_conc.empty:
        type_conc.columns = ["Security Type", "Positions", "Value (USD)", "Avg Duration", "Avg YTM", "% of Portfolio"]
        _add_dataframe_to_sheet(ws_type, type_conc)

    # By Payment Rank sheet
    ws_rank = wb.create_sheet("By Payment Rank")
    rank_conc = get_segmentation_by_payment_rank(db_path)
    if not rank_conc.empty:
        rank_conc.columns = ["Payment Rank", "Positions", "Value (USD)", "Avg Duration", "Avg YTM", "% of Portfolio"]
        _add_dataframe_to_sheet(ws_rank, rank_conc)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return len(alerts)


def generate_port_export(
    output_path: Path,
    db_path: Optional[Path] = None
) -> int:
    """
    Generate Bloomberg PORT upload file (CSV format).

    Returns the number of positions exported.
    """
    positions = get_latest_positions(db_path)

    if positions.empty:
        print("No positions to export.")
        return 0

    # PORT format columns
    port_data = []
    for _, pos in positions.iterrows():
        port_data.append({
            "Security": pos["isin"],  # ISIN as identifier
            "Quantity": pos["face_value"],  # Face value for bonds
            "Price": pos.get("market_price", ""),
            "Cost": "",  # Cost basis if available
            "Currency": pos["currency"],
        })

    port_df = pd.DataFrame(port_data)

    # Save as CSV
    output_path.parent.mkdir(parents=True, exist_ok=True)
    port_df.to_csv(output_path, index=False)

    return len(positions)


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage:")
        print("  python export.py consolidation <output.xlsx>")
        print("  python export.py concentration <output.xlsx>")
        print("  python export.py port <output.csv>")
        sys.exit(1)

    report_type = sys.argv[1]
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None

    if report_type == "consolidation":
        output_path = output_path or Path("consolidation.xlsx")
        count = generate_consolidation_report(output_path)
        print(f"Generated consolidation report with {count} positions: {output_path}")

    elif report_type == "concentration":
        output_path = output_path or Path("concentration.xlsx")
        count = generate_concentration_report(output_path)
        print(f"Generated concentration report with {count} alerts: {output_path}")

    elif report_type == "port":
        output_path = output_path or Path("port_upload.csv")
        count = generate_port_export(output_path)
        print(f"Generated PORT export with {count} positions: {output_path}")

    else:
        print(f"Unknown report type: {report_type}")
        sys.exit(1)
