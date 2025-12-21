"""Bloomberg enrichment workflow - generate request files and import responses."""

import pandas as pd
from pathlib import Path
from datetime import date
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .database import (
    get_securities_needing_enrichment,
    upsert_security,
    mark_securities_enriched,
    get_connection
)


# Bloomberg fields to request
BLOOMBERG_FIELDS = {
    "SECURITY_TYP": "security_type",
    "PAYMENT_RANK": "payment_rank",
    "YLD_YTM_MID": "yield_to_maturity",
    "DUR_ADJ_MID": "modified_duration",
    "RTG_SP": "rating_sp",
    "RTG_MOODY": "rating_moody",
    "CNTRY_OF_RISK": "issuer_country",
    "INDUSTRY_SECTOR": "issuer_sector",
}


def generate_enrichment_request(
    output_path: Path,
    db_path: Optional[Path] = None
) -> int:
    """
    Generate an Excel file with ISINs needing Bloomberg enrichment.

    The Excel file contains BDP formulas that will populate when opened
    in Bloomberg Terminal with Excel Add-in.

    Returns the count of ISINs included.
    """
    # Get securities needing enrichment
    securities = get_securities_needing_enrichment(db_path)

    if securities.empty:
        print("No securities need enrichment.")
        return 0

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Bloomberg Enrichment"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Headers
    headers = ["ISIN", "Name", "Issuer", "Currency"] + list(BLOOMBERG_FIELDS.keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Data rows with BDP formulas
    for row_idx, (_, sec) in enumerate(securities.iterrows(), 2):
        isin = sec["isin"]

        # Static columns
        ws.cell(row=row_idx, column=1, value=isin).border = border
        ws.cell(row=row_idx, column=2, value=sec.get("name", "")).border = border
        ws.cell(row=row_idx, column=3, value=sec.get("issuer", "")).border = border
        ws.cell(row=row_idx, column=4, value=sec.get("currency", "")).border = border

        # BDP formula columns
        for col_idx, field in enumerate(BLOOMBERG_FIELDS.keys(), 5):
            # Create BDP formula: =BDP("ISIN", "FIELD")
            formula = f'=BDP("{isin} Corp", "{field}")'
            cell = ws.cell(row=row_idx, column=col_idx, value=formula)
            cell.border = border

    # Adjust column widths
    column_widths = [14, 40, 30, 8] + [15] * len(BLOOMBERG_FIELDS)
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"[0:2]].width = width

    # Add instructions sheet
    ws_inst = wb.create_sheet("Instructions")
    instructions = [
        "Bloomberg Enrichment Request",
        "",
        "Instructions:",
        "1. Open this file in Excel on a Bloomberg Terminal workstation",
        "2. The BDP formulas will automatically fetch data from Bloomberg",
        "3. Wait for all formulas to calculate (may take a few minutes)",
        "4. Save the file (the formulas will be replaced with values)",
        "5. Use 'enrich-import' command to import the enriched data",
        "",
        "Notes:",
        "- The BDP formulas use 'ISIN Corp' format for corporate bonds",
        "- If a security is not found, try 'ISIN Govt' for government bonds",
        "- Some fields may return #N/A if not available in Bloomberg",
        "",
        f"Generated: {date.today().isoformat()}",
        f"Securities: {len(securities)}",
    ]
    for row, text in enumerate(instructions, 1):
        ws_inst.cell(row=row, column=1, value=text)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return len(securities)


def import_enrichment_response(
    input_path: Path,
    db_path: Optional[Path] = None
) -> tuple[int, int]:
    """
    Import enriched data from Bloomberg Excel file.

    Returns tuple of (updated_count, error_count).
    """
    # Read Excel file
    df = pd.read_excel(input_path, sheet_name=0)

    # Normalize column names
    df.columns = [str(c).strip() for c in df.columns]

    updated = 0
    errors = 0

    for _, row in df.iterrows():
        isin = str(row.get("ISIN", "")).strip()
        if not isin or len(isin) != 12:
            errors += 1
            continue

        try:
            # Map Bloomberg fields to database fields
            updates = {}

            for bb_field, db_field in BLOOMBERG_FIELDS.items():
                value = row.get(bb_field)

                # Skip N/A values
                if pd.isna(value) or str(value).startswith("#"):
                    continue

                # Clean and convert value
                if db_field in ("yield_to_maturity", "modified_duration"):
                    try:
                        updates[db_field] = float(value)
                    except (ValueError, TypeError):
                        pass
                elif db_field == "issuer_country":
                    # Country codes should be 2 chars
                    country = str(value).strip()[:2].upper()
                    if len(country) == 2:
                        updates[db_field] = country
                else:
                    # String fields
                    updates[db_field] = str(value).strip()

            # Normalize security type and payment rank
            if "security_type" in updates:
                updates["security_type"] = _normalize_security_type(updates["security_type"])
            if "payment_rank" in updates:
                updates["payment_rank"] = _normalize_payment_rank(updates["payment_rank"])

            # Update database
            if updates:
                upsert_security(isin, **updates, db_path=db_path)
                updated += 1

        except Exception as e:
            print(f"Error processing {isin}: {e}")
            errors += 1

    # Mark as enriched
    if updated > 0:
        isins = df["ISIN"].dropna().tolist()
        mark_securities_enriched([str(i).strip() for i in isins if len(str(i).strip()) == 12], db_path)

    return updated, errors


def _normalize_security_type(value: str) -> str:
    """Normalize security type to standard values."""
    value = value.upper().strip()

    # Map common Bloomberg values
    mappings = {
        "CORP": "CORPORATE",
        "GOVT": "GOVERNMENT",
        "MUNI": "MUNICIPAL",
        "AT1": "AT1",
        "TIER1": "AT1",
        "TIER2": "TIER2",
        "T2": "TIER2",
        "SUB": "SUBORDINATED",
        "SUBORDINATED": "SUBORDINATED",
        "SENIOR": "SENIOR",
        "SR": "SENIOR",
        "COVERED": "COVERED",
    }

    for key, normalized in mappings.items():
        if key in value:
            return normalized

    return value


def _normalize_payment_rank(value: str) -> str:
    """Normalize payment rank to standard values."""
    value = value.upper().strip()

    mappings = {
        "SENIOR UNSECURED": "SENIOR",
        "SENIOR SECURED": "SENIOR_SECURED",
        "SR UNSECURED": "SENIOR",
        "SR SECURED": "SENIOR_SECURED",
        "SUBORDINATED": "SUBORDINATED",
        "SUB": "SUBORDINATED",
        "JUNIOR SUB": "JUNIOR_SUBORDINATED",
        "JR SUB": "JUNIOR_SUBORDINATED",
        "AT1": "AT1",
        "TIER 1": "AT1",
        "TIER 2": "TIER2",
        "T2": "TIER2",
    }

    for key, normalized in mappings.items():
        if key in value:
            return normalized

    return value


def get_enrichment_status(db_path: Optional[Path] = None) -> dict:
    """Get enrichment status summary."""
    conn = get_connection(db_path)

    total = conn.execute("SELECT COUNT(*) FROM securities").fetchone()[0]
    enriched = conn.execute(
        "SELECT COUNT(*) FROM securities WHERE last_enriched IS NOT NULL"
    ).fetchone()[0]
    needs_enrichment = conn.execute(
        "SELECT COUNT(*) FROM securities WHERE last_enriched IS NULL OR security_type IS NULL"
    ).fetchone()[0]

    # By security type
    by_type = conn.execute("""
        SELECT
            COALESCE(security_type, 'Unknown') as type,
            COUNT(*) as count
        FROM securities
        GROUP BY security_type
        ORDER BY count DESC
    """).df()

    conn.close()

    return {
        "total_securities": total,
        "enriched": enriched,
        "needs_enrichment": needs_enrichment,
        "by_security_type": by_type.to_dict("records")
    }


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage:")
        print("  python enrichment.py request <output.xlsx>  - Generate enrichment request")
        print("  python enrichment.py import <input.xlsx>    - Import enrichment response")
        print("  python enrichment.py status                 - Show enrichment status")
        sys.exit(1)

    command = sys.argv[1]

    if command == "request":
        output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("enrichment_request.xlsx")
        count = generate_enrichment_request(output_path)
        print(f"Generated enrichment request with {count} securities: {output_path}")

    elif command == "import":
        if len(sys.argv) < 3:
            print("Error: Please provide input file path")
            sys.exit(1)
        input_path = Path(sys.argv[2])
        updated, errors = import_enrichment_response(input_path)
        print(f"Imported enrichment data: {updated} updated, {errors} errors")

    elif command == "status":
        status = get_enrichment_status()
        print(f"Total securities: {status['total_securities']}")
        print(f"Enriched: {status['enriched']}")
        print(f"Needs enrichment: {status['needs_enrichment']}")
        print("\nBy security type:")
        for item in status["by_security_type"]:
            print(f"  {item['type']}: {item['count']}")
